#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl"]
# ///
"""Generate a bug bash .xlsx from a BUG-BASH.md file.

Usage: uv run generate_xlsx.py BUG-BASH.md
Parses the `## Groups` section (each `### Group:` becomes a labeled block) and
the `## Ad-hoc / Additional` section (blank fill-in rows).
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# List validations import as Google Sheets dropdown chips; assign chip colors
# once in Sheets after upload (Data > Data validation).
GROUP_COLUMNS = ["Task/Test Case", "What to Validate", "Reporter", "Status", "Notes"]
STATUS_COL = 4  # 1-based index within GROUP_COLUMNS
STATUS_OPTIONS = "Not Started,Pass,Fail"
STATUS_DEFAULT = "Not Started"

ADHOC_COLUMNS = ["Type", "What It Is", "Reporter"]
ADHOC_TYPE_COL = 1
ADHOC_TYPE_OPTIONS = "Bug,Suggested Improvement,Nice to Have,Other"
ADHOC_BLANK_ROWS = 12

# Column widths are per sheet column (both sections share columns A onward).
WIDTHS = [40, 60, 16, 16, 30]


def parse(md: str):
    """Return (title, [(group_name, [(task, validate), ...]), ...], has_adhoc)."""
    title_m = re.search(r"^#\s+(.+)$", md, re.MULTILINE)
    title = title_m.group(1).strip() if title_m else "Bug Bash"

    # Isolate the Groups section (between "## Groups" and the next "## ").
    groups_m = re.search(r"^##\s+Groups\s*$(.*?)(?=^##\s|\Z)", md, re.MULTILINE | re.DOTALL)
    groups_body = groups_m.group(1) if groups_m else ""

    groups = []
    for block in re.split(r"^###\s+Group:\s*", groups_body, flags=re.MULTILINE)[1:]:
        lines = block.splitlines()
        name = lines[0].strip()
        rows = []
        for line in lines[1:]:
            line = line.strip()
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if not cells or re.match(r"^-+$", cells[0].replace(":", "")):
                continue  # separator row
            if cells[0].lower() in ("task", "task/test case"):
                continue  # header row
            task = cells[0]
            validate = cells[1] if len(cells) > 1 else ""
            if task:
                rows.append((task, validate))
        groups.append((name, rows))

    has_adhoc = re.search(r"^##\s+(Ad-?hoc|Additional)", md, re.MULTILINE) is not None
    return title, groups, has_adhoc


def build(title, groups, has_adhoc, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bug Bash"

    group_fill = PatternFill("solid", fgColor="4472C4")
    group_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    status_dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    type_dv = DataValidation(type="list", formula1=f'"{ADHOC_TYPE_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(type_dv)

    r = 1

    def write_section(name, columns, rows, blank_rows=0, dropdown=None):
        """dropdown = (col_index, DataValidation, default_or_None) applied to every body row."""
        nonlocal r
        # Section label row (spans the section's columns).
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns))
        cell = ws.cell(row=r, column=1, value=name)
        cell.fill = group_fill
        cell.font = group_font
        r += 1
        # Column header row.
        for c, col in enumerate(columns, start=1):
            hc = ws.cell(row=r, column=c, value=col)
            hc.fill = header_fill
            hc.font = header_font
        r += 1

        def apply_dropdown(row):
            if dropdown:
                col, dv, default = dropdown
                dv.add(ws.cell(row=row, column=col, value=default))

        # Data rows (text goes in the first two columns).
        for values in rows:
            for c, val in enumerate(values[:2], start=1):
                ws.cell(row=r, column=c, value=val).alignment = wrap
            apply_dropdown(r)
            r += 1
        for _ in range(blank_rows):
            apply_dropdown(r)
            r += 1
        r += 1  # blank separator row between sections

    for name, rows in groups:
        write_section(name, GROUP_COLUMNS, rows, dropdown=(STATUS_COL, status_dv, STATUS_DEFAULT))

    if has_adhoc:
        write_section("Additional Issues", ADHOC_COLUMNS, [], blank_rows=ADHOC_BLANK_ROWS,
                      dropdown=(ADHOC_TYPE_COL, type_dv, None))

    for c, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A1"

    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: generate_xlsx.py BUG-BASH.md")
    md_path = Path(sys.argv[1])
    md = md_path.read_text()
    title, groups, has_adhoc = parse(md)

    slug = re.sub(r"[^a-z0-9]+", "-", title.lower().replace("bug bash:", "")).strip("-")
    out = md_path.with_name(f"{slug or 'bug-bash'}-bug-bash.xlsx")
    build(title, groups, has_adhoc, out)

    total = sum(len(rows) for _, rows in groups)
    print(f"Wrote {out} — {len(groups)} groups, {total} tasks" + (", + ad-hoc section" if has_adhoc else ""))


if __name__ == "__main__":
    main()
